"""韓国代行EMSファイル（Excel）のパーサー
ファイル名: EMS発送リスト_YYMMDD_HHMM.xlsx / EMS발송리스트_YYMMDD_HHMM.xlsx
シート: Sheet1

行1-3: 送付先住所情報（スキップ）
行4  : ヘッダー行
行5以降: データ行（B列が 'Wata' で始まる行）

現在の帳票は日本語ヘッダー版（購入日/発送日/購入番号/EMS番号）なので、
ヘッダー名を優先して動的に列マッピングする。
（旧韓国語ヘッダーにもフォールバック対応）
"""
from datetime import date


class EmsExcelParser:

    SHEET_NAME = 'Sheet1'
    SKIP_ROWS  = 4   # 行1-4はアドレス情報・ヘッダー行

    COL_PURCHASE_DATE = 1   # B: 購入日
    COL_SHIPPED_AT    = 2   # C: 発送日（YYMMDD形式）
    COL_PURCHASE_NO   = 5   # F: 購入番号
    COL_SHOP_NAME     = 7   # H: 仕入先名
    COL_PRODUCT_NAME  = 8   # I: 商品内容
    COL_DANIEL_ID     = 9   # J: ダニエル商品ID
    COL_PRODUCT_CODE  = 10  # K: にゃんたろうずID
    COL_QUANTITY      = 11  # L: 注文数量
    COL_MEMO          = 15  # P: 備考
    COL_EMS_NUMBER    = 16  # Q: EMS番号

    def __init__(self):
        pass

    # ──────────────────────────────────────────────
    def parse(self, file_path):
        """Excelファイルをパースして EMS アイテムリストを返す"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError('openpyxl が必要です: pip install openpyxl')

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if self.SHEET_NAME in wb.sheetnames:
            ws = wb[self.SHEET_NAME]
        else:
            ws = wb.active

        header_map = self._build_header_map(ws)
        items = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < self.SKIP_ROWS:
                continue

            cells = list(row)
            if not any(cells):
                continue

            # B列が 'Wata' で始まる行のみ取り込む
            purchase_date = self._cell(cells, self._idx(header_map, 'purchase_date', self.COL_PURCHASE_DATE))
            if not purchase_date or not purchase_date.startswith('Wata'):
                continue

            # EMS番号（R列）は必須
            ems_raw = self._cell(cells, self._idx(header_map, 'ems_number', self.COL_EMS_NUMBER))
            ems_number = ems_raw.replace(' ', '').strip() if ems_raw else ''
            if not ems_number:
                continue

            # 商品コード: にゃんたろうずIDを優先。空欄/記号時はダニエルIDへフォールバック。
            product_code = self._cell(cells, self._idx(header_map, 'product_code', self.COL_PRODUCT_CODE))
            if not product_code or product_code in ('-', 'ー'):
                product_code = self._cell(cells, self._idx(header_map, 'daniel_id', self.COL_DANIEL_ID))

            item = {
                'ems_number':    ems_number,
                'shipped_at':    self._parse_yymmdd(self._cell(cells, self._idx(header_map, 'shipped_at', self.COL_SHIPPED_AT))),
                'purchase_date': purchase_date,                          # B列: 구매日 作成日
                'purchase_no':   self._cell(cells, self._idx(header_map, 'purchase_no', self.COL_PURCHASE_NO)),
                'order_id':      '',
                'shop_name':     self._cell(cells, self._idx(header_map, 'shop_name', self.COL_SHOP_NAME)),
                'product_name':  self._cell(cells, self._idx(header_map, 'product_name', self.COL_PRODUCT_NAME)),
                'product_code':  product_code,
                'quantity':      self._to_int(cells, self._idx(header_map, 'quantity', self.COL_QUANTITY), default=1),
                'memo':          self._cell(cells, self._idx(header_map, 'memo', self.COL_MEMO)),
            }
            items.append(item)

        wb.close()
        return items

    # ──────────────────────────────────────────────
    @staticmethod
    def _cell(cells, idx):
        if idx >= len(cells):
            return ''
        v = cells[idx]
        return str(v).strip() if v is not None else ''

    @staticmethod
    def _to_int(cells, idx, default=0):
        try:
            v = cells[idx] if idx < len(cells) else None
            return int(float(v)) if v else default
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _parse_yymmdd(raw):
        """
        '250204' → date(2025, 2, 4)
        '260108' → date(2026, 1, 8)
        失敗・空欄時は None を返す（today() デフォルトにしない）
        """
        try:
            s = str(raw).strip()
            # 数字6桁のみ抽出
            digits = ''.join(c for c in s if c.isdigit())[:6]
            if len(digits) == 6:
                yy = int(digits[0:2])
                mm = int(digits[2:4])
                dd = int(digits[4:6])
                return date(2000 + yy, mm, dd)
        except (ValueError, IndexError):
            pass
        return None   # ← today() デフォルトを廃止（sort汚染防止）

    @staticmethod
    def _normalize_header_name(name):
        if not name:
            return ''
        return str(name).replace('\n', '').replace(' ', '').strip().lower()

    def _build_header_map(self, ws):
        """4行目のヘッダーから列名→index を作る（見つからない場合は空）。"""
        header_map = {}
        try:
            header_row = next(ws.iter_rows(min_row=self.SKIP_ROWS, max_row=self.SKIP_ROWS, values_only=True), None)
            if not header_row:
                return header_map
            alias = {
                'purchase_date': ('購入日', '구매일'),
                'shipped_at': ('発送日', '발송일'),
                'purchase_no': ('購入番号', '구매번호'),
                'shop_name': ('仕入先名', '업체명'),
                'product_name': ('商品内容', '상품내용'),
                'daniel_id': ('ダニエル商品id', '상품id'),
                'product_code': ('にゃんたろうずid', '냔타로즈id'),
                'quantity': ('注文数量', '주문수량'),
                'memo': ('備考', '특이사항'),
                'ems_number': ('ems番号', '배송1', 'ems'),
            }
            normalized_headers = [self._normalize_header_name(v) for v in header_row]
            for key, names in alias.items():
                for idx, header in enumerate(normalized_headers):
                    if header in tuple(self._normalize_header_name(n) for n in names):
                        header_map[key] = idx
                        break
        except Exception:
            return {}
        return header_map

    @staticmethod
    def _idx(header_map, key, default_idx):
        return header_map.get(key, default_idx)
