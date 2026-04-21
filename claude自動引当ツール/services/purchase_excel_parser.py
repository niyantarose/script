"""発注リスト Excel (Watanabe_list_*.xlsm) のパーサー
シート: IncomingList
行1-5: ヘッダー行（スキップ）
行6以降: データ行

実際の列構造（ファイル確認済み）:
  A(0) : 구매일/発注日      → 発注日（Wata260414 形式）
  B(1) : 구매번호/PurchaseNo → 発注NO（Wata240801_01 形式）
  C(2) : 주문번호           → Yahoo受注番号
  D(3) : 업체명/VendorName  → 発注先
  E(4) : 상품내용/Product   → 商品名
  F(5) : 상품ID/danielID   → 商品コード（danielID）
  G(6) : 냔타로즈ID         → 内部コード（フォールバック）
  H(7) : 주문수량/Quantity  → 数量
  I(8) : 특이사항/Notes    → メモ

データ行の識別:
  A列が 'Wata' で始まり、かつ B列も 'Wata' で始まる行のみ取込
  （A列のみ 'Wata' の行はセクションヘッダー行のためスキップ）
"""
from datetime import date


class PurchaseExcelParser:

    SHEET_NAME = 'IncomingList'
    SKIP_ROWS  = 5   # 行1-5はヘッダー行

    COL_ORDERED_AT    = 0   # A: 発注日（Wata形式）
    COL_PURCHASE_NO   = 1   # B: 発注NO
    COL_ORDER_ID      = 2   # C: Yahoo受注番号
    COL_SHOP_NAME     = 3   # D: 発注先
    COL_PRODUCT_NAME  = 4   # E: 商品名
    COL_PRODUCT_CODE  = 5   # F: 商品コード（danielID）
    COL_NIYANTAR_CODE = 6   # G: 内部コード（フォールバック）
    COL_QUANTITY      = 7   # H: 数量
    COL_MEMO          = 8   # I: メモ

    def __init__(self):
        pass

    def parse(self, file_path):
        """Excelファイルをパースして発注データのリストを返す"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError('openpyxl が必要です: pip install openpyxl')

        wb = openpyxl.load_workbook(file_path, read_only=True,
                                    data_only=True, keep_vba=False)

        if self.SHEET_NAME in wb.sheetnames:
            ws = wb[self.SHEET_NAME]
        else:
            ws = wb.active

        items = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < self.SKIP_ROWS:
                continue

            cells = list(row)
            if not any(cells):
                continue

            # A列が 'Wata' で始まる行のみ（発注日の確認）
            raw_date = self._cell(cells, self.COL_ORDERED_AT)
            if not raw_date or not str(raw_date).startswith('Wata'):
                continue

            # B列も 'Wata' で始まる行のみ（セクションヘッダーを除外）
            purchase_no = self._cell(cells, self.COL_PURCHASE_NO)
            if not purchase_no or not str(purchase_no).startswith('Wata'):
                continue

            # 商品コード: F列優先 → 空なら G列
            product_code = self._cell(cells, self.COL_PRODUCT_CODE)
            if not product_code:
                product_code = self._cell(cells, self.COL_NIYANTAR_CODE)
            if not product_code:
                continue

            item = {
                'ordered_at':   self._parse_wata_date(raw_date),
                'purchase_no':  purchase_no,
                'order_id':     self._cell(cells, self.COL_ORDER_ID),
                'shop_name':    self._cell(cells, self.COL_SHOP_NAME),
                'product_name': self._cell(cells, self.COL_PRODUCT_NAME),
                'product_code': product_code,
                'quantity':     self._to_int(cells, self.COL_QUANTITY, default=1),
                'memo':         self._cell(cells, self.COL_MEMO),
            }
            items.append(item)

        wb.close()
        return items

    @staticmethod
    def _cell(cells, idx):
        if idx >= len(cells):
            return ''
        v = cells[idx]
        return str(v).strip() if v is not None else ''

    @staticmethod
    def _to_int(cells, idx, default=0):
        try:
            v = cells[idx] if idx < len(cells) else None
            return int(float(v)) if v else default
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _parse_wata_date(raw):
        """
        'Wata260414' → date(2026, 4, 14)
        'Wata240801' → date(2024, 8, 1)
        失敗時は今日の日付を返す
        """
        try:
            s = str(raw).strip()
            if s.startswith('Wata') and len(s) >= 10:
                yy = int(s[4:6])
                mm = int(s[6:8])
                dd = int(s[8:10])
                return date(2000 + yy, mm, dd)
        except (ValueError, IndexError):
            pass
        return date.today()
